import sys
import sqlite3
import copy
import string
import operator
import openpyxl
from helper import preprint, postprint, notvalid, line
import re 

class Connection:
    '''Connect to the local db.'''    
    def __init__(self):
        self.conn = sqlite3.connect("example.db")
        self.cursor = self.conn.cursor()
    
    def zavrit(self):
        self.conn.close()


class Dictionary:
    '''Dictionary object with methods to show the list of the words 
    in db, to insert a word to the db and close the connection.
    '''
    
    def __init__(self):
        '''Connects to the db. Creates dictionary and its deepcopy - to 
        compare them for following actions. Finally, closes the 
        connection to the db.
        '''
        self.conn = Connection()
        self.word_d = list(self.conn.cursor.execute("SELECT word_d FROM words"))
        self.word_tuple = tuple(self.conn.cursor.execute("SELECT article, word_c FROM words"))
        self.dictionary = {}
        x = 0
        for i in self.word_d:
            self.dictionary[i[0]] = self.word_tuple[x]
            x += 1
        self.dictionary2 = copy.deepcopy(self.dictionary)
        self.conn.zavrit()
        
    def show(self, **kwargs):
        '''Takes language input for listing and sorting all the words in the 
        dictionary by the given language number.
        '''
        try:
            languaged = ""
            prelanguaged = int(kwargs["prelanguage"])
        except:
            preprint()
            print ("Enter number 1 or 2")
            postprint()
            return False
        if (prelanguaged < 1)  or (prelanguaged > 2):
            preprint()
            print ("Not a valid choice.")
            postprint()
            return False
        prelist_for_display = [(key, self.dictionary[key]) for key in self.dictionary]
        list_for_display = [(i[1][0], i[0], i[1][1]) for i in prelist_for_display]
        if prelanguaged == 1:
            list_for_display.sort(key = operator.itemgetter(2))
            preprint()
            print ("Sorted alphabetically in your language")
            line()
        else:
            preprint()
            print ("Sorted alphabetically in German")
            line()
            list_for_display.sort(key = operator.itemgetter(1))       
        for i in list_for_display:
            if i[0] == "":
                print ("{0:10s} - {1:>13s}" .format(i[1], i[2]))
            else:
                print ("{0:2s} {1:10s} - {2:>13s}" .format(i[0], i[1], i[2]))
        print ("----------------------------------")
        print("Number of words: ", len(self.dictionary))
        print ("\n")
        return list_for_display
    
    def insertion(self, article, word_d, word_c):
        '''Connect to the db. Insert a word to the db. Closes db. '''
        self.conn = Connection()
        inserted = [(article, word_d, word_c)]
        self.conn.cursor.executemany("INSERT INTO words VALUES (?, ?,?)", inserted)
        self.conn.conn.commit()
        self.conn.zavrit()
        print ("\n")
        print ("The word has been added/changed.")
        print ("\n")
    
    def checker(self, *argv):
        '''Checks all given normal words, if they contain correct characters. 
        If not - returns False. If yes - returns True.
        '''
        for arg in argv:
            if arg == None or len(arg) > 50 or arg == "":
                notvalid()
                return False
        not_accepted = string.digits + string.punctuation
        for i in not_accepted:
            for arg in argv:
                if i in arg:
                   notvalid()
                   return False
        return True

    def article_checker(self, article):
        '''Checks articles of a German word, if it contains correct 
        characters  and has the correct length. If not - returns False.
         If yes - returns True.
        '''
        if len(article) > 3 or len(article) == 1 or len(article) == 2 :
                notvalid()
                return False
        else:
            return True
    
    def insert_a_word(self, article, word_d, word_c):
        '''Checks all the words via checker() and updates the first 
        dictionary from the first connection to the db. The original 
        object of the dictionary from the db is updated. If the word 
        is not in the original dictionary, then the word is inserted
        to the db, else the db is not updated.
        '''
        control = self.checker(word_d, word_c) and self.article_checker(article)
        if not control:
            return False
        self.dictionary.update({word_d: (article, word_c)})
        if word_d not in self.dictionary2:
            self.insertion(article, word_d, word_c)
            return [article, word_d, word_c]
        else:
            if self.dictionary[word_d] != self.dictionary2[word_d]:
                self.insertion(article, word_d, word_c)
                return [article, word_d, word_c]
            else:
                print ("\n")
                print ("The word pair:", article, word_d, "-", word_c, 
                    "is already in the dictionary.")
                return False

    def find(self, word):
        '''Finds and prints all the found words. Both languages.'''
        word = word.lower()
        control = self.checker(word)
        if not control:
            return False
        lower_case_list = [i.lower() for i in list(self.dictionary.keys())]
        if word in lower_case_list:
            if len(word) > 1:
                word = word.capitalize()
            preprint()
            print ("{0:2s} {1:10s} - {2:>13s}" 
                .format(self.dictionary[word][0], word, self.dictionary[word][1]))
            postprint()
            print("Number of the found words: 1")
            postprint()
            return [self.dictionary[word][0], word, self.dictionary[word][1]]
        list_of_words_your_language = [e[1] for e in self.dictionary.values()]
        indexes_of_searched_word = []
        for i in list(enumerate(list_of_words_your_language, start=0)):
            if i[1] == word:
                indexes_of_searched_word.append(i[0])
        final = []
        x = 0
        for i in list_of_words_your_language: 
            if i == word:
                keys = list(self.dictionary.keys())
                index = indexes_of_searched_word[x]
                final.append((self.dictionary[keys[index]][0],
                    keys[index],list_of_words_your_language[index]))
                x += 1
        if word in list_of_words_your_language:
            preprint()
            for i in final:
                print ("{0:2s} {1:10s} - {2:>13s}" .format(i[0], i[1], i[2]))
            postprint()
            print("Number of the found words: ", len(final))
            print ("\n")
            return [self.dictionary[keys[index]][0], keys[index], word]
        if word not in self.dictionary or word not in self.dictionary.values():
            print ("\n")
            print (word, " is not in the dictionary")
            postprint()
            print("Number of the found words: 0")
            print ("\n")
            return False
            
        
    def delete(self, deleted_word):
        '''Deletes a German word from the db.'''
        self.conn = Connection()
        postdeleted_word = copy.deepcopy(deleted_word)
        deleted_word = [(deleted_word[1],)]
        self.conn.cursor.executemany("DELETE FROM words WHERE word_d = ?", deleted_word)
        self.conn.conn.commit()
        self.conn.zavrit()
        preprint()
        print ("The word:", postdeleted_word[0], postdeleted_word[1], " - ", postdeleted_word[2])
        postprint()
        print ("The word has been deleted.")
        print("\n")
        del self.dictionary[postdeleted_word[1]]

    def importing(self, file_name, sheet_name, number_of_words):
        '''Imports excel sheet of the .xlsx or .xls formats to the db. 
        Sends single word to insert_a_word() function.
        '''
        try:
            number_of_words = int(number_of_words)
            file_name = str(file_name)
            sheet_name = str(sheet_name)
            if re.match("[a-zA-Z0-9]*\.xlsx$", file_name):
                pass
            else:
                raise Exception()
            wb = openpyxl.load_workbook(file_name)
            sheet = wb[sheet_name]
        except:
            line()
            print("Not valid entries: file must be in the same " 
                "dictionary and the name of it must be written correctly,sheet "
                "must exist with the correct name and the number of the " 
                "rows must be correct.")
            line()
            return False
        for i in range(1, number_of_words + 1, 1):
            self.insert_a_word(sheet.cell(row = i, column = 1).value, 
                sheet.cell(row = i, column = 2).value, 
                sheet.cell(row = i, column = 3).value)

    def export(self, exported_file, sheet_name):
        '''Exports the dictionary to an excel sheet in .xlsx format.'''
        try:
            exported_file = str(exported_file)
            sheet_name = str(sheet_name)
        except:
            line()
            print("Not valid entries, try again.")
            line()
            return False
        wb = openpyxl.Workbook()
        wb.create_sheet(index=0, title = sheet_name)
        sheet = wb[sheet_name]
        k = 0
        for i in self.dictionary:
            k += 1
            sheet["A" + str(k)] = self.dictionary[i][0]
            sheet["B" + str(k)] = i
            sheet["C" + str(k)] = self.dictionary[i][1]
        exported_file = exported_file + ".xlsx"
        wb.save(exported_file)
        preprint()
        print ("The dictionary was exported to the file of the app named: "
            , exported_file)
        postprint()

class Menu:
    '''Menu handling of the app in the command prompt.'''    
    
    def __init__(self):
        self.seznamek = Dictionary()
    
    def run_it(self, choice):
        '''Menu driving function'''
        try:
            choosen = ""
            choosen = int(choice)
        except:
            print ("\n")
            print ("Enter valid number choice.")
            print ("\n")
            return False
        choice = int(choice)
        choices = {
                1: self.process_it,
                2: self.seznam_list,
                3: self.find_it,
                4: self.delete_it,
                5: self.import_it,
                6: self.export_it,
                7: self.quity,
                }
        if int(choice) > len(choices) or int(choice) <= 0:
            print ("\n")
            print ("Not a valid choice.")
            print ("\n")
        else:
            return choices[choice]
        
    def display(self):
        '''Displays menu.'''
        print("\n")
        print("MENU: ")
        line()
        print("1. Insert/change a word")
        print("2. List all the words in the dictionary")
        print("3. Find a word in the dictonary both in German or your language")
        print("4. Delete a word from the dictionary")
        print("5. Import an .xlsx file of the words in the following " 
            "format: 1st column article or empty, 2nd German word, 3rd " 
            "translation of the German word")
        print("6. Export the dictionary to .xlsx excel file")
        print("7. Quit the dictionary")
        postprint()
        self.prompt_it()
    
    def prompt_it(self):
        '''Input runs functions displayed in tha main menu.'''
        choice = input ("Enter your choice: ")
        action = self.run_it(choice)
        if action:
            action()
        
    def process_it(self):
        '''Sending inputs to inserting function.'''
        article = input("Article of the German word: ")
        word_d = input("German word: ")
        word_c = input("Word in your language: ")
        self.seznamek.insert_a_word(article, word_d, word_c)
        
    def seznam_list(self):
        '''Initiating display of the words in the dictionary.'''
        prelanguage = input("For alphabet sorting in your language " 
            "enter 1, for alphabet sorting in German, enter 2: ")
        return (self.seznamek.show(prelanguage=prelanguage))
    
    def find_it(self):
        '''Initiating search function for the input word.'''
        word = input("Find this word, do not use articles with German words: ")
        return (self.seznamek.find(word))
    
    def delete_it(self):
        '''Initiating delete function for the input word.'''
        word = input("Delete this word, do not use articles with German words: ")
        deleted_word = self.seznamek.find(word)
        if deleted_word == False:
            return
        else:
            self.seznamek.delete(deleted_word)

    def import_it(self):
        '''Sends inputs to importing(), which imports excel sheet of the
         .xlsx or .xls formats to the db. Sends single word to 
         insert_a_word function.
        '''
        file_name = input("What is the name of the file: ")
        sheet_name = input("What is the name of the imported sheet of "
            "the German - other language word pairs: ")
        number_of_words = input("How many of the German - other language" 
            "word pairs are in the excel sheet: ")
        return self.seznamek.importing(file_name, sheet_name, number_of_words)

    def export_it(self):
        '''Sends inputs to export(), which exports excel sheet of the 
        .xlsx or .xls formats from the db.
        '''
        exported_file = input("What name of the exported file do you want?"
            " Please, don't include the .xlsx in the file name: ")
        sheet_name = input("What is the name of the sheet for the export: ")
        return self.seznamek.export(exported_file, sheet_name)

    def quity(self):
        '''Quits the dictionary.'''
        sys.exit()
        
    def run(self):
        '''Runs the dictionary and its choices.'''
        while True:
            self.display()


if __name__ == "__main__":
    Menu().run()